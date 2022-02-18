from myClasses import *
import openpyxl
from openpyxl import load_workbook

import mdToArray
import os
import shutil

from openpyxl.styles import Font


class ArrayToExcel:
    def __init__(self):
        self.book = None
        self.templateBookPath=None
        self.templateSheetName = None
        self.startRow = None
        self.startCol = None

        self.baseFont=None

        return

    def setBook(self, book: Book):
        self.book = book

    def readTemplate(self, path, templateSheetName: str, startRow: int, startCol: int):
        # self.wb = load_workbook(filename=path, keep_vba=True)
        self.templateBookPath=path
        self.templateSheetName = templateSheetName
        self.startCol = startCol
        self.startRow = startRow

    def generateBook(self,outputPath:str,font:str,size):
        
        outputDir=outputPath[0:outputPath.rfind("\\")]
        if not os.path.exists(outputPath):
            if not os.path.exists(outputDir):
                os.makedirs(outputDir)
            shutil.copy(self.templateBookPath, outputPath)
            # wb=load_workbook(filename=self.templateBookPath, keep_vba=True,read_only=False)

        wb=load_workbook(outputPath,keep_vba=True,read_only=False)

        for sheet in self.book.sheets:  # type:Sheet
            print(sheet.sheetName)
            print(sheet.data)
            
            #delete before created sheet
            if sheet.sheetName in wb.sheetnames:
                std=wb.get_sheet_by_name(sheet.sheetName)
                wb.remove_sheet(std)
                
            
            ws = wb.copy_worksheet(
                wb.get_sheet_by_name(self.templateSheetName))
            ws.title = sheet.sheetName
            # rootFont=ws.cell(self.startRow+1, self.startCol+1).font
            # self.baseFont=Font(name=rootFont.name,sz=rootFont.sz)
            self.baseFont=Font(name=font,size=size)

            for r, row in enumerate(sheet.data):
                for c, column in enumerate(row):
                    if column != "":
                        self.__setVal(ws, r+1, c+1, column)
        
        wb.save(outputPath)

        wb.close()

        return

    def __setVal(self, ws: openpyxl.worksheet, row, col, val):
        cell=ws.cell(row=row+self.startRow, column=col+self.startCol)
        cell.font=self.baseFont
        cell.value=val
        return


if __name__ == "__main__":
    # mte = mdToArray.MdToArray()
    # mte.read("mdDocs/sample.md")
    # mte.compile()
    import pickle
    # with open("book.pickle","wb")as f:
    #     pickle.dump(mte.book, f)
    with open("book.pickle", "rb")as f:
        book = pickle.load(f)
    ate = ArrayToExcel()
    ate.setBook(book=book)
    ate.readTemplate("format.xlsm", "template", 3, 3)
    ate.generateBook()
