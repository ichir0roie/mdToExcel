import openpyxl
from openpyxl import load_workbook

from myClasses import *

import os, os.path
import win32com.client

bookName="combineBooks.xlsm"

class CombineBooks:
    def __init__(self):
        self.stng=Settings("settings.json")

        return
    
    def setInfo(self):
        wb=openpyxl.load_workbook(bookName,keep_vba=True,read_only=False)
        ws=wb["settings"]
        
        ws["C2"]=self.stng.outputExFolder
        ws["C3"]=self.stng.combineFolder
        ws["C5"]=self.stng.ignoreSheetNames

        wb.save(bookName)
        wb.close()
    
    def runVBA(self):
        if os.path.exists(bookName):
            xl=win32com.client.Dispatch("Excel.Application")
            xl.Workbooks.Open(os.path.abspath(bookName), ReadOnly=1)
            xl.Application.Run(bookName+"!ThisWorkbook.mergeFiles")
            ##    xl.Application.Save() # if you want to save then uncomment this line and change delete the ", ReadOnly=1" part from the open function.
            xl.Application.Quit() # Comment this out if your excel script closes
        del xl

    def run(self):
        self.setInfo()
        self.runVBA()

if __name__=="__main__":
    cb=CombineBooks()
    cb.setInfo()
    cb.runVBA()

        